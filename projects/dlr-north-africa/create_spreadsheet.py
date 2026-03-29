#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF", size=10)
header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
subheader_fill = PatternFill(start_color="E94560", end_color="E94560", fill_type="solid")
subheader_font = Font(bold=True, color="FFFFFF", size=10)
data_font = Font(size=10)
bold_font = Font(bold=True, size=10)
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
green_fill = PatternFill(start_color="E8F5E9", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF8E1", fill_type="solid")
red_fill = PatternFill(start_color="FCE4EC", fill_type="solid")

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap
        cell.border = thin_border

def style_subheader(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = wrap
        cell.border = thin_border

def style_data(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font
        cell.alignment = wrap
        cell.border = thin_border

def auto_width(ws, cols, min_w=12, max_w=35):
    for c in range(1, cols+1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(min_w, 18), max_w)

# ============ SHEET 1: Comparative Summary ============
ws1 = wb.active
ws1.title = "Summary"
headers = ["Metric", "Morocco", "Algeria", "Tunisia"]
data = [
    ["Central Bank", "Bank Al-Maghrib (BAM)", "Banque d'Algérie", "Banque Centrale de Tunisie (BCT)"],
    ["Governor", "Abdellatif Jouahri (since 2003)", "Mouatassem Boudiaf (Jan 2026)", "Fethi Zouhair Nouri (Feb 2024)"],
    ["Director General / Deputy", "Abderrahim Bouazza (DG)", "Vice-Governor TBC", "Vice-Governor TBC"],
    ["Currency", "MAD (Dirham)", "DZD (Dinar)", "TND (Dinar)"],
    ["ISO Code", "MAD", "DZD", "TND"],
    ["Population", "37.5 million", "45.6 million", "12.4 million"],
    ["GDP", "$150 billion", "$240 billion", "$48 billion"],
    ["Inflation", "~2%", "~4%", "~7-8%"],
    ["FX Reserves", "~$36 billion", "$64.6 billion", "~$8B (3.5 months imports)"],
    ["Exchange Regime", "Managed (±5% band, 60% EUR / 40% USD)", "Managed float", "Managed float (closed currency)"],
    ["Denominations", "20, 50, 100, 200 MAD", "200, 500, 1000, 2000 DZD", "5, 10, 20, 50 TND"],
    ["Highest Denomination", "200 MAD (~$20)", "2000 DZD (~$15)", "50 TND (~$16)"],
    ["Substrate", "Paper (Durasafe® for commemorative)", "Paper", "Paper + varnish"],
    ["Primary Printer", "Dar As-Sikkah (domestic)", "Domestic + external", "Crane Currency"],
    ["DLR Historical Link", "Pre-1987 printer", "Component supply", "1972 series printer"],
    ["Cash Economy Level", "Moderate", "Very high (50%+ outside banks)", "High (closed currency)"],
    ["Last New Series", "2023", "2022", "2022"],
    ["Current Supplier Relationship", "Dar As-Sikkah self-sufficient", "Domestic + unknown external", "Crane Currency (RAPID® HD)"],
    ["DLR Priority", "MEDIUM (components)", "HIGH (volume + new governor)", "MEDIUM-LOW (Crane incumbent)"],
    ["Key Contact Point", "Governor Jouahri / DG Bouazza", "Governor Boudiaf (new)", "Governor Nouri"],
]

ws1.append(headers)
style_header(ws1, 1, 4)
for i, row in enumerate(data):
    ws1.append(row)
    style_data(ws1, i+2, 4)
    ws1.cell(row=i+2, column=1).font = bold_font
    if "HIGH" in str(row[-1]) or "HIGH" in str(row[1]) or "HIGH" in str(row[2]):
        for c in range(1,5):
            if "HIGH" in str(ws1.cell(row=i+2, column=c).value or ""):
                ws1.cell(row=i+2, column=c).fill = green_fill

auto_width(ws1, 4, 20, 40)

# ============ SHEET 2: Morocco Denominations ============
ws2 = wb.create_sheet("Morocco Banknotes")
headers = ["Denomination", "Colour", "Size (mm)", "Substrate", "Printer", "Key Feature", "Series Year", "Status"]
data = [
    ["20 MAD", "Purple/blue", "138 × 69", "Paper", "Dar As-Sikkah", "King Mohammed VI", "2012", "Circulating"],
    ["50 MAD", "Green", "147 × 70", "Paper", "Dar As-Sikkah", "King Mohammed VI, Bab Essebaa", "2012", "Circulating"],
    ["100 MAD", "Brown", "153 × 72", "Paper", "Dar As-Sikkah", "Three Kings (redesigned)", "2023", "Circulating"],
    ["200 MAD", "Blue", "159 × 74", "Paper", "Dar As-Sikkah", "King Mohammed VI", "2012", "Circulating"],
    ["25 MAD (Comm.)", "—", "—", "Durasafe® hybrid", "Dar As-Sikkah", "25th anniversary DAS", "2012", "Commemorative"],
    ["50 MAD (Comm.)", "—", "—", "Paper", "Dar As-Sikkah", "BAM 50th anniversary", "2009", "Commemorative"],
]
ws2.append(headers)
style_header(ws2, 1, 8)
for i, row in enumerate(data):
    ws2.append(row)
    style_data(ws2, i+2, 8)
auto_width(ws2, 8)

# ============ SHEET 3: Algeria Denominations ============
ws3 = wb.create_sheet("Algeria Banknotes")
headers = ["Denomination", "Colour", "Series", "Year Issued", "Theme/Design", "Security Features", "Status"]
data = [
    ["200 DA", "Reddish brown", "4th series", "1992/1996", "Koranic motifs, mosque", "Standard", "Circulating"],
    ["500 DA", "Violet/pink", "4th series", "1992/1998", "Numidian battle", "Holographic strip (1998+)", "Circulating"],
    ["500 DA", "Violet/pink", "New series", "2018", "Alcomsat-1 satellite", "Enhanced security", "Circulating"],
    ["1,000 DA", "Red/brown", "4th series", "1992/1998", "Tassili n'Ajjer prehistory", "Holographic strip (1998+)", "Circulating"],
    ["1,000 DA", "Blue/red", "New series", "2018", "Grand Mosque of Algiers", "Enhanced security", "Circulating"],
    ["2,000 DA", "Purple/green", "—", "2011", "University, science, research", "Standard", "Circulating"],
    ["2,000 DA", "Red/blue/green", "—", "2020/2021", "FLN historical leaders", "Enhanced security", "Circulating"],
    ["2,000 DA", "Green", "—", "2022", "Martyrs' Memorial, Great Mosque", "Enhanced security", "Circulating"],
]
ws3.append(headers)
style_header(ws3, 1, 7)
for i, row in enumerate(data):
    ws3.append(row)
    style_data(ws3, i+2, 7)
auto_width(ws3, 7)

# ============ SHEET 4: Tunisia Denominations ============
ws4 = wb.create_sheet("Tunisia Banknotes")
headers = ["Denomination", "Colour", "Size (mm)", "Year", "Portrait", "Security Features", "Printer", "Status"]
data = [
    ["5 TND", "Green", "143 × 73", "2022", "Slaheddine El Amami", "RAPID® HD thread, Spark® Live", "Crane Currency", "Circulating"],
    ["10 TND", "Blue/yellow", "148 × 73", "2020", "Dr. Tawhida Ben Cheikh", "RAPID® HD thread", "Crane Currency", "Circulating"],
    ["20 TND", "Red/orange", "153 × 76", "2017", "Farhat Hached", "RAPID® thread", "Crane Currency", "Circulating"],
    ["50 TND", "Brown", "158 × 79", "2022", "Hedi Nouira", "RAPID® HD thread, Spark® Live", "Crane Currency", "Circulating"],
]
ws4.append(headers)
style_header(ws4, 1, 8)
for i, row in enumerate(data):
    ws4.append(row)
    style_data(ws4, i+2, 8)
auto_width(ws4, 8)

# ============ SHEET 5: Org Chart Data ============
ws5 = wb.create_sheet("Key Personnel")
headers = ["Country", "Institution", "Name", "Title/Role", "Since", "Notes"]
data = [
    ["Morocco", "Bank Al-Maghrib", "Abdellatif Jouahri", "Governor", "2003", "Longest-serving, 23 years"],
    ["Morocco", "Bank Al-Maghrib", "Abderrahim Bouazza", "Director General", "—", "Deputises for Governor"],
    ["Morocco", "Bank Al-Maghrib", "Fathallah Oualalou", "Board Member", "—", ""],
    ["Morocco", "Bank Al-Maghrib", "Mustapha Moussaoui", "Board Member / Audit Chair", "—", ""],
    ["Morocco", "Bank Al-Maghrib", "Larbi Jaïdi", "Board Member", "—", ""],
    ["Morocco", "Bank Al-Maghrib", "Najat El Mekkaoui", "Board Member / Social Funds Chair", "—", ""],
    ["Morocco", "Bank Al-Maghrib", "Mohammed Daïri", "Board Member", "—", ""],
    ["Morocco", "Bank Al-Maghrib", "Mouna Cherkaoui", "Board Member / Audit Committee", "—", ""],
    ["Morocco", "Bank Al-Maghrib", "Mohammed Tarik Bchir", "Board Member", "—", ""],
    ["Morocco", "Dar As-Sikkah", "—", "State Printing Works Director", "Est. 1987", "Domestic banknote printer, Rabat"],
    ["Algeria", "Banque d'Algérie", "Mouatassem Boudiaf", "Governor", "Jan 2026", "Newly appointed — predecessor dismissed"],
    ["Algeria", "Banque d'Algérie", "TBC", "Vice-Governor", "—", "Frequent leadership changes"],
    ["Algeria", "Banque d'Algérie", "Salah Eddine Taleb", "Former Governor", "May 2022 – Jan 2026", "Dismissed Jan 2026"],
    ["Tunisia", "Banque Centrale de Tunisie", "Fethi Zouhair Nouri", "Governor", "Feb 2024", "Economist, professor since 1982, 6-year term"],
    ["Tunisia", "Banque Centrale de Tunisie", "TBC", "Vice-Governor", "—", "6-year term, presidential decree"],
    ["Tunisia", "Banque Centrale de Tunisie", "Marouane El Abassi", "Former Governor", "2018-2024", "Predecessor"],
]
ws5.append(headers)
style_header(ws5, 1, 6)
for i, row in enumerate(data):
    ws5.append(row)
    style_data(ws5, i+2, 6)
    # Colour by country
    if row[0] == "Morocco":
        ws5.cell(row=i+2, column=1).fill = PatternFill(start_color="C8E6C9", fill_type="solid")
    elif row[0] == "Algeria":
        ws5.cell(row=i+2, column=1).fill = PatternFill(start_color="BBDEFB", fill_type="solid")
    elif row[0] == "Tunisia":
        ws5.cell(row=i+2, column=1).fill = PatternFill(start_color="FFE0B2", fill_type="solid")
auto_width(ws5, 6, 15, 40)

# ============ SHEET 6: Competitive Landscape ============
ws6 = wb.create_sheet("Competitors")
headers = ["Competitor", "Morocco", "Algeria", "Tunisia", "Notes"]
data = [
    ["Dar As-Sikkah (domestic)", "PRIMARY PRINTER", "—", "—", "Morocco's state printing works, est. 1987"],
    ["Bank of Algeria (domestic)", "—", "Has printing capability", "—", "Capacity may be limited"],
    ["Crane Currency", "—", "Unknown", "CURRENT SUPPLIER", "RAPID® HD proprietary technology"],
    ["Oberthur Fiduciaire", "Possible substrate", "Possible", "Possible", "French-origin, 70+ countries"],
    ["Giesecke+Devrient", "Possible components", "Possible", "Unknown", "German, major competitor"],
    ["SICPA", "Ink supply", "Ink supply", "Ink supply", "Swiss, security inks"],
    ["De La Rue", "Historical (pre-1987)", "Component opportunity", "Historical (1972 series)", "Re-entry opportunities exist"],
]
ws6.append(headers)
style_header(ws6, 1, 5)
for i, row in enumerate(data):
    ws6.append(row)
    style_data(ws6, i+2, 5)
auto_width(ws6, 5, 18, 40)

# Save
output = "/home/ubuntu/clawd/projects/dlr-north-africa/DLR_North_Africa_Data.xlsx"
wb.save(output)
print(f"Saved: {output}")
