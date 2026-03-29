#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(bold=True, color="FFFFFF", size=10)
header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
data_font = Font(size=10)
bold_font = Font(bold=True, size=10)
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
green_fill = PatternFill(start_color="C8E6C9", fill_type="solid")
blue_fill = PatternFill(start_color="BBDEFB", fill_type="solid")
orange_fill = PatternFill(start_color="FFE0B2", fill_type="solid")

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = wrap; cell.border = thin_border

def style_data(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font; cell.alignment = wrap; cell.border = thin_border

def auto_width(ws, cols, min_w=14, max_w=38):
    for c in range(1, cols+1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(min_w, 18), max_w)

# SHEET 1: Summary
ws1 = wb.active
ws1.title = "Summary"
headers = ["Metric", "Nigeria", "Seychelles", "Zimbabwe"]
data = [
    ["Central Bank", "Central Bank of Nigeria (CBN)", "Central Bank of Seychelles (CBS)", "Reserve Bank of Zimbabwe (RBZ)"],
    ["Governor", "Olayemi Cardoso (Sep 2023)", "Caroline Abel", "Dr. John Mushayavanhu (Mar 2024)"],
    ["Deputy/DG", "4 Deputy Governors", "Brian Commettant (1st DG), Mike Tirant (2nd DG)", "TBC"],
    ["Currency", "NGN (Naira)", "SCR (Rupee)", "ZWG (ZiG — Zimbabwe Gold)"],
    ["ISO Code", "NGN", "SCR", "ZWG"],
    ["Population", "230+ million", "~100,000", "~16.5 million"],
    ["GDP", "$475 billion", "$2.1 billion", "$30 billion"],
    ["Inflation", "~29%", "~2.3%", "3.8% (Feb 2026)"],
    ["FX Reserves", "$46.7 billion", "Moderate", "$1.2 billion (gold + FX backing ZiG)"],
    ["Exchange Rate", "~₦1,600/USD (floating)", "~14 SCR/USD", "~26 ZiG/USD"],
    ["Denominations", "₦5, ₦10, ₦20, ₦50, ₦100, ₦200, ₦500, ₦1000", "R25, R50, R100, R500", "ZiG 10, 20, 50, 100, 200"],
    ["Substrate", "Polymer (₦5-50) + Paper (₦100-1000)", "Paper", "Paper"],
    ["Primary Printer", "NSPM (CBN 90%, DLR 1%)", "De La Rue ✅", "Domestic (Fidelity Printers?)"],
    ["DLR Relationship", "1% NSPM stake + 2022 designer", "Sole printer (Gemini™ tech)", "Unknown — opportunity"],
    ["Last New Series", "2022 redesign (DLR designed)", "2016 Biodiversity series (DLR)", "BiG 5 series — April 2026"],
    ["Cash Economy", "Very high (large unbanked pop.)", "Moderate (tourism USD/EUR)", "70-85% USD transactions"],
    ["DLR Priority", "HIGHEST (volume + existing link)", "RETENTION (existing client)", "HIGH (active opportunity)"],
    ["Key Contact", "Gov. Cardoso / Dir. Adedeji (Currency Ops)", "Gov. Caroline Abel / DG Mike Tirant", "Gov. Dr. Mushayavanhu"],
]
ws1.append(headers)
style_header(ws1, 1, 4)
for i, row in enumerate(data):
    ws1.append(row)
    style_data(ws1, i+2, 4)
    ws1.cell(row=i+2, column=1).font = bold_font
auto_width(ws1, 4, 20, 42)

# SHEET 2: Nigeria Banknotes
ws2 = wb.create_sheet("Nigeria Banknotes")
headers = ["Denomination", "Colour", "Size (mm)", "Substrate", "Status", "Key Features", "Series"]
data = [
    ["₦5", "Green/grey", "130 × 72", "Polymer", "Circulating", "Alhaji Aliyu Mai-Bornu", "2009+"],
    ["₦10", "Grey/red", "130 × 72", "Polymer", "Circulating", "Alvan Ikoku", "2009+"],
    ["₦20", "Green", "130 × 72", "Polymer", "Circulating", "Gen. Murtala Mohammed", "2007+"],
    ["₦50", "Blue", "130 × 72", "Polymer", "Circulating", "Various national figures", "2009+"],
    ["₦100", "Brown/purple", "151 × 78", "Paper", "Circulating", "Chief Obafemi Awolowo, Zuma Rock", "1999/2014"],
    ["₦200", "Green", "151 × 78", "Paper", "Circulating (old + 2022 redesign)", "Sir Ahmadu Bello", "2000/2022"],
    ["₦500", "Blue/grey", "151 × 78", "Paper", "Circulating (old + 2022 redesign)", "Dr. Nnamdi Azikiwe", "2001/2022"],
    ["₦1,000", "Red/green", "151 × 78", "Paper", "Circulating (old + 2022 redesign)", "Mai-Bornu & Isong. Holographic strip.", "2005/2022"],
    ["₦100 (Comm.)", "Special", "—", "Paper", "Commemorative", "Centennial of Nigeria (2014)", "2014"],
]
ws2.append(headers)
style_header(ws2, 1, 7)
for i, row in enumerate(data):
    ws2.append(row)
    style_data(ws2, i+2, 7)
auto_width(ws2, 7)

# SHEET 3: Seychelles Banknotes
ws3 = wb.create_sheet("Seychelles Banknotes")
headers = ["Denomination", "Theme", "Substrate", "Printer", "Security Features", "Series", "Status"]
data = [
    ["R25", "Seychelles Biodiversity", "Paper", "De La Rue", "Gemini™ UV, holographic patch", "2016", "Circulating"],
    ["R50", "Aldabra rail (flightless bird)", "Paper", "De La Rue", "Holographic sailfish→50, fluorescent thread 2.5mm", "2016 (upgraded 2011)", "Circulating"],
    ["R100", "Seychelles giant tortoise", "Paper", "De La Rue", "Gold holographic sailfish→100, colour-shifting thread 2.5mm", "2016 (upgraded 2011)", "Circulating"],
    ["R500", "Seychelles scops owl", "Paper", "De La Rue", "Gold holographic sailfish→500, colour-shifting thread 3mm", "2016 (upgraded 2011)", "Circulating"],
    ["R10 (discontinued)", "Converted to coin", "—", "—", "—", "2016", "Now a coin"],
]
ws3.append(headers)
style_header(ws3, 1, 7)
for i, row in enumerate(data):
    ws3.append(row)
    style_data(ws3, i+2, 7)
auto_width(ws3, 7)

# SHEET 4: Zimbabwe Banknotes
ws4 = wb.create_sheet("Zimbabwe Banknotes")
headers = ["Denomination", "Series", "Colour", "Design Front", "Design Reverse", "Security Features", "Release Date", "Status"]
data = [
    ["ZiG 10", "Original 2024", "Navy blue", "Domboremari rocks", "Gold bars", "Basic security, QR code", "30 Apr 2024", "Being withdrawn"],
    ["ZiG 20", "Original 2024", "Peach/green", "Domboremari rocks", "Gold bars", "Basic security, QR code", "13 May 2024", "Being withdrawn"],
    ["ZiG 10", "BiG 5 (2026)", "TBC", "Buffalo + RBZ logo", "Matopo Hills", "Intaglio, magnetic thread, PEAK, latent image, colour-shift", "7 Apr 2026", "NEW — Launching"],
    ["ZiG 20", "BiG 5 (2026)", "Brown", "Elephant + RBZ logo", "Parliament Building", "Intaglio, magnetic thread, PEAK, latent image, colour-shift", "7 Apr 2026", "NEW — Launching"],
    ["ZiG 50", "BiG 5 (2026)", "TBC", "Rhinoceros + RBZ logo", "TBC", "Intaglio, magnetic thread, PEAK, latent image", "7 Apr 2026", "NEW — Launching"],
    ["ZiG 100", "BiG 5 (2026)", "TBC", "Leopard + RBZ logo", "TBC", "Rolling Star thread, colour-shift gold→green", "Later 2026", "Pending"],
    ["ZiG 200", "BiG 5 (2026)", "TBC", "Lion + RBZ logo", "TBC", "Intaglio, colour-shifting gold bars", "Later 2026", "Pending"],
]
ws4.append(headers)
style_header(ws4, 1, 8)
for i, row in enumerate(data):
    ws4.append(row)
    style_data(ws4, i+2, 8)
auto_width(ws4, 8)

# SHEET 5: Key Personnel
ws5 = wb.create_sheet("Key Personnel")
headers = ["Country", "Institution", "Name", "Title/Role", "Since", "Notes"]
data = [
    ["Nigeria", "CBN", "Olayemi Cardoso", "Governor", "Sep 2023", "Appointed by President Tinubu. Transparency/reform agenda."],
    ["Nigeria", "CBN", "TBC", "DG Financial System Stability", "—", "One of 4 Deputy Governors"],
    ["Nigeria", "CBN", "TBC", "DG Corporate Services", "—", ""],
    ["Nigeria", "CBN", "TBC", "DG Economic Policy", "—", ""],
    ["Nigeria", "CBN", "TBC", "DG Operations", "—", ""],
    ["Nigeria", "CBN", "Adetona Sikiru Adedeji", "Director of Currency Operations", "2025", "Current — signature on 2025 notes"],
    ["Nigeria", "CBN", "Mohammed-Jamiu O. Solaja", "Former Dir. Currency Operations", "2024", "Signature on 2024 notes"],
    ["Nigeria", "NSPM", "Ahmed Halilu", "Managing Director", "2018+", "Suggested DLR for 2022 redesign"],
    ["Seychelles", "CBS", "Caroline Abel", "Governor & Chairperson", "—", ""],
    ["Seychelles", "CBS", "Brian Commettant", "First Deputy Governor", "—", ""],
    ["Seychelles", "CBS", "Mike Tirant", "Second Deputy Governor", "—", "Head of Banking Services"],
    ["Seychelles", "CBS", "Prof. William Ogara", "Non-Executive Director", "—", ""],
    ["Seychelles", "CBS", "Dr. Sherley Marie", "Non-Executive Director", "—", ""],
    ["Seychelles", "CBS", "Mr. James Jean", "Non-Executive Director", "—", ""],
    ["Seychelles", "CBS", "Mr. Jean-Paul Barbier", "Non-Executive Director", "—", ""],
    ["Zimbabwe", "RBZ", "Dr. John Mushayavanhu", "Governor", "Mar 2024", "Launched ZiG. Business ties to President."],
    ["Zimbabwe", "Govt", "Prof. Mthuli Ncube", "Minister of Finance", "—", "Works with RBZ on monetary policy"],
    ["Zimbabwe", "RBZ", "John Mangudya", "Former Governor", "2014-2024", "Now CEO Mutapa Investment Fund"],
]
ws5.append(headers)
style_header(ws5, 1, 6)
for i, row in enumerate(data):
    ws5.append(row)
    style_data(ws5, i+2, 6)
    if row[0] == "Nigeria": ws5.cell(row=i+2, column=1).fill = green_fill
    elif row[0] == "Seychelles": ws5.cell(row=i+2, column=1).fill = blue_fill
    elif row[0] == "Zimbabwe": ws5.cell(row=i+2, column=1).fill = orange_fill
auto_width(ws5, 6, 15, 42)

# SHEET 6: Competitive Landscape
ws6 = wb.create_sheet("Competitors")
headers = ["Competitor", "Nigeria", "Seychelles", "Zimbabwe", "Notes"]
data = [
    ["De La Rue", "1% NSPM stake + 2022 designer", "SOLE PRINTER ✅", "Unknown — opportunity", "Historical relationships across all three"],
    ["NSPM (domestic)", "PRIMARY PRINTER (CBN 90%)", "—", "—", "DLR 1% stake, capacity limited"],
    ["Fidelity Printers (domestic)", "—", "—", "Likely printer", "Linked to RBZ/Fidelity Gold Refinery"],
    ["Crane Currency", "Unknown", "—", "Unknown", "Competitor — RAPID® technology"],
    ["Oberthur Fiduciaire", "Possible", "—", "Possible", "French-origin printer"],
    ["Giesecke+Devrient", "Possible", "—", "Possible", "German, major competitor"],
    ["SICPA", "Ink supply possible", "—", "Ink supply possible", "Swiss, security inks"],
]
ws6.append(headers)
style_header(ws6, 1, 5)
for i, row in enumerate(data):
    ws6.append(row)
    style_data(ws6, i+2, 5)
auto_width(ws6, 5, 18, 42)

output = "/home/ubuntu/clawd/projects/dlr-sub-saharan/DLR_Nigeria_Seychelles_Zimbabwe_Data.xlsx"
wb.save(output)
print(f"Saved: {output}")
