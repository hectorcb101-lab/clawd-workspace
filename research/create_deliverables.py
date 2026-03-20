#!/usr/bin/env python3
"""Create quantum computing research deliverables: Excel spreadsheet and PDF report."""

import subprocess
import sys

# Ensure dependencies
for pkg in ['openpyxl', 'markdown', 'weasyprint']:
    try:
        __import__(pkg)
    except ImportError:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '-q'])

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ============================================================
# SPREADSHEET
# ============================================================
wb = Workbook()

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_align = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_sheet(ws, headers, data):
    ws.append(headers)
    for row in data:
        ws.append(row)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = cell_align
            cell.border = thin_border
    # Auto-size
    for col_idx in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 45)

# Sheet 1: Key Players
ws1 = wb.active
ws1.title = "Key Players"
style_sheet(ws1, 
    ["Company", "Type", "Physical Qubits", "Logical Qubits", "Key Achievement", "Funding", "HQ"],
    [
        ["Google Quantum AI", "Superconducting", "65 (Willow)", "Below-threshold QEC", "Quantum Echoes: 13,000× speedup, verifiable quantum advantage", "Internal (Alphabet)", "Santa Barbara, CA, US"],
        ["IBM Quantum", "Superconducting", "156+ (2,299 total)", "Roadmap 2026", "HSBC bond trading 34% improvement; 3.6T+ circuits; 97% uptime", "Internal (IBM)", "Yorktown Heights, NY, US"],
        ["Microsoft", "Topological", "~8 (Majorana 1)", "0 (not demonstrated)", "First topological qubit claim (Feb 2025, Nature); scepticism remains", "Internal (Microsoft)", "Redmond, WA, US"],
        ["IonQ", "Trapped Ion (Yb)", "64", "Roadmap 2027-28", "99.99% 2-qubit gate fidelity world record; first QC co. >$100M revenue", "$3.6B+ total", "College Park, MD, US"],
        ["Quantinuum", "Trapped Ion (Ba)", "98 (Helios)", "48-94 demonstrated", "2:1 encoding ratio; beyond break-even QEC; DARPA QBI Stage B", "$900M+ ($10B+ val)", "Cambridge, UK"],
        ["PsiQuantum", "Photonic", "N/A (photonic)", "Targeting million-scale", "Omega chipset (Nature, Feb 2025); GlobalFoundries fab partnership", "$600M+ cumulative", "Palo Alto, CA, US"],
        ["QuEra Computing", "Neutral Atom", "3,000 (array)", "96 demonstrated", "3,000-qubit array sustained 2+ hours; Harvard/MIT collaboration", "$230M+ (2025)", "Boston, MA, US"],
        ["Atom Computing", "Neutral Atom", "1,000+", "Partnership w/ Microsoft", "Microsoft partnership for neutral-atom + software stack", "Undisclosed", "Boulder, CO, US"],
        ["Infleqtion", "Neutral Atom", "100 (Sqale)", "N/A", "UK's only operational 100-qubit system at NQCC (Mar 2026)", "NYSE: INFQ", "Boulder, CO / Oxford, UK"],
        ["Pasqal", "Neutral Atom", "100+", "N/A", "Leading EU neutral-atom company; French national champion", "€100M+", "Massy, France"],
        ["Rigetti", "Superconducting", "84+", "N/A", "Two system purchase orders ($5.7M); cloud-native approach", "$300M+", "Berkeley, CA, US"],
        ["Origin Quantum", "Superconducting", "100+", "N/A", "China's leading QC startup; state-backed", "State-funded", "Hefei, China"],
        ["SpinQ", "Various", "Various", "N/A", "Desktop and portable quantum computers for education", "Private", "Shenzhen, China"],
    ]
)

# Sheet 2: Timeline of Breakthroughs
ws2 = wb.create_sheet("Timeline of Breakthroughs")
style_sheet(ws2,
    ["Date", "Event", "Company/Institution", "Significance"],
    [
        ["Aug 2024", "NIST finalises 3 PQC standards (ML-KEM, ML-DSA, SLH-DSA)", "NIST", "First post-quantum cryptography standards ready for implementation"],
        ["Oct 2024", "Jiuzhang 3.0 photonic quantum advantage", "USTC (China)", "Solved sampling problem in 1μs vs 20B years classically"],
        ["Dec 2024", "Google Willow chip: below-threshold QEC", "Google", "First demonstration that more qubits = fewer errors in surface code"],
        ["Feb 2025", "Microsoft Majorana 1 topological qubit", "Microsoft", "First claimed topological qubit; published in Nature; scepticism remains"],
        ["Feb 2025", "PsiQuantum Omega photonic chipset", "PsiQuantum", "Published in Nature; all components for million-qubit photonic QC"],
        ["May 2025", "AdaBoost.Q quantum ML on superconducting processor", "Nature/npj QI", "Quantum ensemble learning demonstrated on real hardware"],
        ["Jun 2025", "US Executive Order: PQC federal migration", "White House", "Mandated government-wide migration to quantum-resistant cryptography"],
        ["Jun 2025", "IBM Nighthawk processor roadmap update", "IBM", "Higher-connectivity processor for complex circuits"],
        ["Jul 2025", "EU Quantum Europe Strategy launched", "European Commission", "Roadmap to make EU 'quantum industrial powerhouse' by 2030"],
        ["Sep 2025", "HSBC bond trading quantum trial: 34% improvement", "HSBC + IBM", "First empirical evidence of QC value in financial services"],
        ["Oct 2025", "IonQ 99.99% two-qubit gate fidelity world record", "IonQ", "10× error reduction; underpins 256-qubit 2026 and million-qubit 2030 roadmap"],
        ["Oct 2025", "Google Quantum Echoes: 13,000× speedup", "Google", "First verifiable quantum advantage algorithm on hardware"],
        ["Nov 2025", "Quantinuum Helios: 48 logical qubits, 2:1 ratio", "Quantinuum", "Best encoding ratio; 98 physical → 48 logical qubits"],
        ["2025", "QuEra 3,000-qubit array, 96 logical qubits", "QuEra", "Largest qubit array; most logical qubits demonstrated"],
        ["2025", "IonQ exceeds $100M annual revenue", "IonQ", "First pure-play QC company to pass $100M revenue milestone"],
        ["2025", "Global QC investment reaches 75% of 2024 total in 5 months", "Industry-wide", "Fewer rounds but much larger cheques; $50M+ average"],
        ["Jan 2026", "Quantinuum S-1 IPO filing, ~$20B target", "Quantinuum", "Largest QC IPO attempt; validates trapped-ion approach"],
        ["Feb 2026", "Nature: long-lived remote ion-ion entanglement", "USTC / Nature", "Scalable quantum repeaters; 1.2M Bell pairs at 11km"],
        ["Mar 2026", "China 15th Five-Year Plan targets quantum leadership", "Chinese Government", "1,000+ qubit control, space-earth quantum network, AI+quantum"],
        ["Mar 2026", "Quantinuum: 94 logical qubits, beyond break-even QEC", "Quantinuum", "Encoded circuits outperform unencoded; logical gate errors ~1:10,000"],
        ["Mar 2026", "Infleqtion delivers UK's 100-qubit system to NQCC", "Infleqtion", "UK's only operational 100-qubit quantum computer"],
    ]
)

# Sheet 3: National Strategies
ws3 = wb.create_sheet("National Strategies")
style_sheet(ws3,
    ["Country/Region", "Investment ($)", "Focus Areas", "Key Programmes", "Key Strengths"],
    [
        ["United States", "$5B+ (public) + dominant private capital", "Hardware, software, QEC, defence", "DARPA QBI, National Quantum Initiative, NSF Quantum Leap", "Private sector depth (IonQ, Google, IBM, PsiQuantum); DARPA pipeline"],
        ["China", "$15B+ (public, estimated)", "Quantum communication, computing, sensing", "15th Five-Year Plan, USTC labs, space-earth quantum network", "State coordination; world's largest QC network (12,000km); Pan Jianwei team"],
        ["European Union", "€11B+ (public, 5-year)", "Research, infrastructure, skills, space/dual-use", "Quantum Flagship, Quantum Europe Strategy (Jul 2025)", "Largest researcher/startup pool; strong science; weak commercialisation"],
        ["United Kingdom", "£2.5B (10-year strategy)", "QEC, trapped ions, neutral atoms, PQC", "National Quantum Strategy, NQCC", "Quantinuum HQ; Oxford/Cambridge ecosystem; Riverlane; Infleqtion NQCC"],
        ["Japan", "$1.5B+ (estimated)", "Superconducting, quantum annealing", "Quantum Innovation Strategy, Riken QC Centre", "D-Wave heritage; Fujitsu quantum-inspired; strong materials science"],
        ["Canada", "$1B+ (estimated)", "Photonic, superconducting, quantum sensing", "National Quantum Strategy (2023)", "Xanadu, D-Wave; Perimeter Institute; strong academic base"],
        ["Australia", "$A1B+ (announced)", "Silicon qubits, trapped ions", "National Quantum Strategy (2023)", "Silicon Quantum Computing; UNSW research; Five Eyes alignment"],
        ["South Korea", "$2B+ (announced)", "Superconducting, quantum communication", "K-Quantum Initiative", "Samsung, SK involvement; strong semiconductor manufacturing base"],
    ]
)

# Sheet 4: Commercial Applications
ws4 = wb.create_sheet("Commercial Applications")
style_sheet(ws4,
    ["Sector", "Use Case", "Key Companies", "Status (Mar 2026)", "Timeline to Production"],
    [
        ["Finance", "Bond price prediction/optimisation", "HSBC, IBM, JPMorgan, Quantinuum", "Proven (34% improvement, Sep 2025)", "2026-2028"],
        ["Finance", "Portfolio optimisation", "JPMorgan, Quantinuum, Goldman Sachs", "Pilot/Research", "2027-2029"],
        ["Finance", "Risk analysis / Monte Carlo", "Goldman Sachs, IBM", "Research", "2028-2030"],
        ["Pharma", "Drug discovery / molecular simulation", "IBM, Cleveland Clinic, Moderna, Roche", "Research/Early pilot", "2029-2033"],
        ["Materials", "Materials design / catalysis", "Google, IBM, BASF, Quantinuum", "Research", "2028-2032"],
        ["Logistics", "Supply chain optimisation", "D-Wave, Volkswagen, DHL", "Quantum-inspired (classical); QC pilot", "2027-2030"],
        ["Cybersecurity", "PQC migration", "NIST, Google, Cloudflare, all enterprises", "Standards finalised; migration underway", "2024-2030 (mandatory)"],
        ["Cybersecurity", "Quantum key distribution", "QuantumCTek, ID Quantique, Toshiba", "Operational (China 12,000km network)", "Operational now"],
        ["Defence/Aerospace", "Sensing, simulation, cryptanalysis", "Various (classified)", "Active programmes (DARPA QBI)", "2026-2033"],
        ["Energy", "Battery chemistry / grid optimisation", "IBM, Google, utilities", "Research", "2030-2035"],
        ["AI/ML", "Quantum-enhanced ML, optimisation", "Google, IBM, Nvidia (CUDA Quantum)", "Research/hybrid approaches", "2028-2032"],
        ["Certified Randomness", "Random number generation", "Google (Willow), Quantinuum", "Demonstrated", "2025-2026"],
    ]
)

# Sheet 5: Market Data
ws5 = wb.create_sheet("Market Data")
style_sheet(ws5,
    ["Metric", "Value", "Source", "Date"],
    [
        ["Global QC market size 2025", "$1.44 billion", "Precedence Research", "Feb 2026"],
        ["Global QC market size 2025 (alt.)", "$2.01 billion", "InsightAce Analytic", "Jan 2026"],
        ["Global QC market size 2026 (projected)", "~$1.88-2.0 billion", "Multiple sources", "Jan 2026"],
        ["Global QC market size 2035 (projected)", "$19.44 billion", "Precedence Research", "Feb 2026"],
        ["Global QC market size 2035 (alt.)", "$40.45 billion", "InsightAce Analytic", "Jan 2026"],
        ["Quantum 2.0 market 2036 (projected)", "$50 billion", "Research and Markets", "Feb 2026"],
        ["CAGR 2026-2035", "29.7-36.0%", "Multiple sources", "2026"],
        ["QC companies revenue 2024", "$650-750 million", "McKinsey", "2025"],
        ["QC companies revenue 2025 (est.)", "$1 billion+", "McKinsey", "2025"],
        ["IonQ FY2025 revenue", "$100M+ (first to milestone)", "IonQ Earnings", "Feb 2026"],
        ["Quantum computers sold 2024", "37 units", "Resonance", "2025"],
        ["QC hardware sales value 2024", "$854 million", "Resonance", "2025"],
        ["Average QC unit price 2024", "$19 million", "Resonance/QuantumBasel", "2025"],
        ["Average QC unit price 2021", "$48 million", "Resonance/QuantumBasel", "2025"],
        ["Average funding round 2025", "$50M+", "ResearchAndMarkets", "Jul 2025"],
        ["H1 2025 investment", "~75% of full 2024 total", "The Quantum Insider", "Jun 2025"],
        ["Total QC investment 2024", "~$2 billion", "QuantumBasel", "Oct 2025"],
        ["Emerging startups share of 2024 funding", "37%", "McKinsey", "2025"],
        ["IonQ total funding", "$3.6 billion+", "IonQ", "2026"],
        ["QuEra 2025 raise", "$230 million+", "QuEra", "2025"],
        ["Quantinuum valuation (IPO target)", "~$20 billion", "Quantinuum S-1", "Jan 2026"],
        ["Quantinuum total funding", "$900 million+", "Various", "2026"],
        ["EU public quantum investment (5-year)", "€11 billion+", "European Commission", "Jul 2025"],
        ["UK quantum strategy (10-year)", "£2.5 billion", "UK Government", "2023"],
    ]
)

xlsx_path = os.path.expanduser("~/clawd/research/quantum-computing-2026-data.xlsx")
wb.save(xlsx_path)
print(f"Spreadsheet saved: {xlsx_path}")

# ============================================================
# PDF
# ============================================================
import markdown

md_path = os.path.expanduser("~/clawd/obsidian-vault/Research/EX - Quantum Computing Technology 2025-2026 (2026-03-17).md")
with open(md_path, 'r') as f:
    md_content = f.read()

html_body = markdown.markdown(md_content, extensions=['tables', 'fenced_code'])

html_full = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
@page {{ size: A4; margin: 2cm; }}
body {{ font-family: 'Helvetica Neue', Arial, sans-serif; font-size: 11pt; line-height: 1.6; color: #1a1a1a; }}
h1 {{ color: #1F4E79; font-size: 22pt; border-bottom: 3px solid #1F4E79; padding-bottom: 8px; page-break-after: avoid; }}
h2 {{ color: #2E75B6; font-size: 16pt; border-bottom: 1px solid #ccc; padding-bottom: 4px; margin-top: 24px; page-break-after: avoid; }}
h3 {{ color: #375F8C; font-size: 13pt; margin-top: 18px; page-break-after: avoid; }}
table {{ border-collapse: collapse; width: 100%; margin: 12px 0; font-size: 9.5pt; }}
th {{ background-color: #1F4E79; color: white; padding: 8px; text-align: left; }}
td {{ padding: 6px 8px; border: 1px solid #ddd; }}
tr:nth-child(even) {{ background-color: #f5f8fc; }}
a {{ color: #2E75B6; text-decoration: none; }}
strong {{ color: #1a1a1a; }}
blockquote {{ border-left: 4px solid #2E75B6; padding-left: 12px; color: #555; margin: 12px 0; }}
hr {{ border: none; border-top: 2px solid #1F4E79; margin: 24px 0; }}
.header {{ text-align: center; margin-bottom: 30px; }}
.header h1 {{ font-size: 28pt; margin-bottom: 4px; }}
.header p {{ color: #666; font-size: 11pt; }}
</style>
</head>
<body>
<div class="header">
<h1>🔬 Quantum Computing Technology</h1>
<p>Exhaustive Research Report — 2025–2026 Updates</p>
<p>Prepared by Atlas for Finn McKie | 17 March 2026</p>
</div>
<hr>
{html_body}
</body>
</html>"""

html_path = os.path.expanduser("~/clawd/research/quantum-report-temp.html")
with open(html_path, 'w') as f:
    f.write(html_full)

pdf_path = os.path.expanduser("~/clawd/research/quantum-computing-2026-report.pdf")

try:
    from weasyprint import HTML
    HTML(filename=html_path).write_pdf(pdf_path)
    print(f"PDF saved: {pdf_path}")
except Exception as e:
    print(f"WeasyPrint failed: {e}")
    print("Trying wkhtmltopdf fallback...")
    import subprocess
    try:
        subprocess.run(['wkhtmltopdf', '--page-size', 'A4', '--margin-top', '20mm', 
                       '--margin-bottom', '20mm', '--margin-left', '20mm', '--margin-right', '20mm',
                       html_path, pdf_path], check=True, capture_output=True)
        print(f"PDF saved (wkhtmltopdf): {pdf_path}")
    except Exception as e2:
        print(f"wkhtmltopdf also failed: {e2}")
        # Last resort: use pandoc
        try:
            subprocess.run(['pandoc', md_path, '-o', pdf_path, 
                          '--pdf-engine=xelatex', '-V', 'geometry:margin=2cm'], 
                          check=True, capture_output=True)
            print(f"PDF saved (pandoc): {pdf_path}")
        except Exception as e3:
            print(f"All PDF methods failed. HTML saved at: {html_path}")

# Cleanup
if os.path.exists(html_path) and os.path.exists(pdf_path):
    os.remove(html_path)

print("Done!")
